"""
excel_parser.py
---------------
Excel 파일을 파싱하여 스키마, enum, 데이터를 구조화된 형태로 반환한다.

지원 시트:
  _Schema  : 메시지 필드 정의
  _Enum    : enum 타입 정의
  그 외    : 데이터 시트 (시트명 = MessageName)

동작 모드:
  단일 파일 모드  : parse_excel(path)           — 하나의 .xlsx 안에 스키마+데이터
  멀티 파일 모드  : parse_excel_multi(schema_dir, data_dir)
                     - schema_dir: Proto_*.xlsx 파일들 (스키마/enum 정의)
                     - data_dir  : *.xlsx 데이터 파일들 (시트명 = MessageName)

주의: openpyxl은 .xlsx 형식만 지원한다. .xls 파일은 .xlsx로 재저장 필요.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class FieldDef:
    """스키마 시트의 단일 필드 정의."""
    message_name: str
    field_name: str
    field_type: str    # 예: "int32", "string", "enum:PawnType"
    field_number: int
    rule: str          # "optional" 또는 "repeated"
    comment: str


@dataclass
class EnumValueDef:
    """Enum 시트의 단일 값 정의."""
    enum_name: str
    value_name: str
    value_number: int
    comment: str


@dataclass
class ExcelData:
    """파싱된 Excel 전체 데이터 컨테이너."""
    # {MessageName: [FieldDef, ...]}
    schema: dict[str, list[FieldDef]] = field(default_factory=dict)
    # {EnumName: [EnumValueDef, ...]}
    enums: dict[str, list[EnumValueDef]] = field(default_factory=dict)
    # {MessageName: [{"FieldName": raw_value, ...}, ...]}
    data: dict[str, list[dict[str, Any]]] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _cell_str(value: Any) -> str:
    """셀 값을 문자열로 변환하고 공백을 제거한다."""
    if value is None:
        return ""
    return str(value).strip()


def _is_row_empty(row) -> bool:
    """행의 모든 셀이 None이면 True를 반환한다."""
    return all(cell.value is None for cell in row)


def _safe_int(value: Any) -> int:
    """Excel 숫자 셀은 float로 반환될 수 있으므로 안전하게 int로 변환한다."""
    if value is None:
        raise ValueError("필드 번호가 비어 있습니다.")
    return int(float(value))


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _parse_schema(ws) -> dict[str, list[FieldDef]]:
    """
    _Schema 시트를 파싱한다.

    헤더 행: MessageName | FieldName | FieldType | FieldNumber | Rule | Comment
    MessageName 셀이 비어 있으면 이전 행의 값을 이어받는다(carry-forward).
    """
    schema: dict[str, list[FieldDef]] = {}
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return schema

    # 헤더 인덱스 매핑
    header = [_cell_str(c).lower() for c in rows[0]]
    try:
        idx = {
            "message_name": header.index("messagename"),
            "field_name":   header.index("fieldname"),
            "field_type":   header.index("fieldtype"),
            "field_number": header.index("fieldnumber"),
            "rule":         header.index("rule"),
            "comment":      header.index("comment"),
        }
    except ValueError as e:
        raise ValueError(f"_Schema 시트 헤더 오류: {e}") from e

    current_message = ""
    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue

        msg = _cell_str(row[idx["message_name"]])
        if msg:
            current_message = msg
        if not current_message:
            continue

        field_name   = _cell_str(row[idx["field_name"]])
        field_type   = _cell_str(row[idx["field_type"]])
        rule         = _cell_str(row[idx["rule"]]) or "optional"
        comment      = _cell_str(row[idx["comment"]])

        if not field_name or not field_type:
            continue

        try:
            field_number = _safe_int(row[idx["field_number"]])
        except (ValueError, TypeError) as e:
            raise ValueError(
                f"_Schema 시트 {row_num}행: FieldNumber 변환 실패 — {e}"
            ) from e

        if current_message not in schema:
            schema[current_message] = []

        schema[current_message].append(FieldDef(
            message_name=current_message,
            field_name=field_name,
            field_type=field_type,
            field_number=field_number,
            rule=rule,
            comment=comment,
        ))

    return schema


def _parse_enums(ws) -> dict[str, list[EnumValueDef]]:
    """
    _Enum 시트를 파싱한다.

    헤더 행: EnumName | ValueName | ValueNumber | Comment
    EnumName 셀이 비어 있으면 이전 행의 값을 이어받는다(carry-forward).
    """
    enums: dict[str, list[EnumValueDef]] = {}
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return enums

    header = [_cell_str(c).lower() for c in rows[0]]
    try:
        idx = {
            "enum_name":    header.index("enumname"),
            "value_name":   header.index("valuename"),
            "value_number": header.index("valuenumber"),
            "comment":      header.index("comment"),
        }
    except ValueError as e:
        raise ValueError(f"_Enum 시트 헤더 오류: {e}") from e

    current_enum = ""
    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue

        name = _cell_str(row[idx["enum_name"]])
        if name:
            current_enum = name
        if not current_enum:
            continue

        value_name = _cell_str(row[idx["value_name"]])
        comment    = _cell_str(row[idx["comment"]])

        if not value_name:
            continue

        try:
            value_number = _safe_int(row[idx["value_number"]])
        except (ValueError, TypeError) as e:
            raise ValueError(
                f"_Enum 시트 {row_num}행: ValueNumber 변환 실패 — {e}"
            ) from e

        if current_enum not in enums:
            enums[current_enum] = []

        enums[current_enum].append(EnumValueDef(
            enum_name=current_enum,
            value_name=value_name,
            value_number=value_number,
            comment=comment,
        ))

    return enums


def _parse_data_sheet(ws, sheet_name: str) -> list[dict[str, Any]]:
    """
    데이터 시트를 파싱한다.

    1행: FieldName 헤더
    2행~: 실제 데이터 (전체 빈 행 스킵)

    동일 컬럼명이 여러 번 등장하면 각 행의 값을 리스트로 수집한다
    (None 값은 제외). 이 방식으로 repeated 필드를 다중 컬럼으로 입력할 수 있다.
    단일 컬럼은 기존과 동일하게 스칼라 값으로 반환한다.
    """
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return []

    header = [_cell_str(c) for c in rows[0]]

    # 컬럼명 → 인덱스 목록 매핑 (중복 컬럼명 지원)
    col_indices: dict[str, list[int]] = {}
    for i, name in enumerate(header):
        if not name:
            continue
        col_indices.setdefault(name, []).append(i)

    repeated_cols = {name for name, indices in col_indices.items() if len(indices) > 1}

    data: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue

        row_dict: dict[str, Any] = {}
        for col_name, indices in col_indices.items():
            if col_name in repeated_cols:
                # 여러 컬럼의 값을 리스트로 수집 (None은 제외)
                values = [row[i] for i in indices if row[i] is not None]
                row_dict[col_name] = values if values else None
            else:
                row_dict[col_name] = row[indices[0]]

        if row_dict:
            data.append(row_dict)

    return data


# ---------------------------------------------------------------------------
# Public API — 단일 파일 모드
# ---------------------------------------------------------------------------

def parse_excel(path: Path) -> ExcelData:
    """
    Excel 파일을 파싱하여 ExcelData를 반환한다.

    Args:
        path: .xlsx 파일 경로

    Returns:
        ExcelData: schema, enums, data를 담은 컨테이너

    Raises:
        FileNotFoundError: 파일이 존재하지 않을 때
        ValueError: 시트 구조가 올바르지 않을 때
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"Excel 파일을 찾을 수 없습니다: {path}\n"
            "힌트: openpyxl은 .xlsx 형식만 지원합니다. "
            ".xls 파일은 Excel에서 '다른 이름으로 저장 → .xlsx'로 변환하세요."
        )

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        raise ValueError(
            f"Excel 파일을 열 수 없습니다: {path}\n"
            f"오류: {e}\n"
            "힌트: openpyxl은 .xlsx 형식만 지원합니다."
        ) from e

    result = ExcelData()

    # _Schema 시트 파싱
    if "_Schema" not in wb.sheetnames:
        raise ValueError("_Schema 시트를 찾을 수 없습니다.")
    result.schema = _parse_schema(wb["_Schema"])

    # _Enum 시트 파싱 (없어도 허용)
    if "_Enum" in wb.sheetnames:
        result.enums = _parse_enums(wb["_Enum"])

    # 데이터 시트 파싱
    reserved = {"_Schema", "_Enum"}
    for sheet_name in wb.sheetnames:
        if sheet_name in reserved:
            continue
        result.data[sheet_name] = _parse_data_sheet(wb[sheet_name], sheet_name)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Public API — 멀티 파일 모드 (B안)
# ---------------------------------------------------------------------------

def _open_workbook(path: Path):
    """파일 존재 여부를 확인하고 workbook을 반환한다."""
    if not path.exists():
        raise FileNotFoundError(
            f"Excel 파일을 찾을 수 없습니다: {path}\n"
            "힌트: openpyxl은 .xlsx 형식만 지원합니다."
        )
    try:
        return openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        raise ValueError(f"Excel 파일을 열 수 없습니다: {path}\n오류: {e}") from e


_ENUM_FILENAME = "Enum.xlsx"


def _merge_enums(
    target: dict[str, list],
    new_enums: dict[str, list],
    source_name: str,
) -> None:
    """
    new_enums를 target에 병합한다. 중복 EnumName이 있으면 오류를 발생시킨다.

    Args:
        target:      병합 대상 dict (in-place 수정)
        new_enums:   병합할 새 enum dict
        source_name: 오류 메시지에 표시할 파일명
    """
    for enum_name, values in new_enums.items():
        if enum_name in target:
            raise ValueError(
                f"EnumName '{enum_name}'이 중복 정의되어 있습니다.\n"
                f"  파일: {source_name}\n"
                "  동일한 Enum은 하나의 파일에만 정의하세요."
            )
        target[enum_name] = values


def parse_schema_dir(schema_dir: Path) -> tuple[dict, dict]:
    """
    schema_dir 내의 파일들을 파싱하고 스키마와 enum을 병합한다.

    파싱 순서 및 파일 역할:
      1. Enum.xlsx (있으면)
         - 특정 테이블에 종속되지 않는 공용 enum 정의
         - _Enum 시트만 읽음
      2. Proto_*.xlsx
         - 테이블별 메시지 스키마 정의
         - _Schema 시트(필수) + _Enum 시트(선택, 해당 테이블 전용 enum)

    서로 다른 파일에서 동일한 MessageName 또는 EnumName이 중복 정의되면 오류를 발생시킨다.

    Args:
        schema_dir: 스키마 파일들이 위치하는 디렉터리

    Returns:
        (merged_schema, merged_enums) 튜플

    Raises:
        FileNotFoundError: schema_dir이 없거나 Proto_*.xlsx 파일이 없을 때
        ValueError: MessageName 또는 EnumName 중복 시
    """
    schema_dir = Path(schema_dir)
    if not schema_dir.exists():
        raise FileNotFoundError(f"스키마 디렉터리를 찾을 수 없습니다: {schema_dir}")

    merged_schema: dict[str, list] = {}
    merged_enums:  dict[str, list] = {}

    # ── 1단계: Enum.xlsx 파싱 (없어도 무관) ────────────────────────────
    enum_file = schema_dir / _ENUM_FILENAME
    if enum_file.exists():
        print(f"  공용 Enum 파일 파싱: {enum_file.name}")
        wb = _open_workbook(enum_file)
        if "_Enum" not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"{enum_file.name}: _Enum 시트가 없습니다. "
                "Enum.xlsx에는 반드시 _Enum 시트가 있어야 합니다."
            )
        _merge_enums(merged_enums, _parse_enums(wb["_Enum"]), enum_file.name)
        wb.close()
    else:
        print(f"  Enum.xlsx 없음 — 공용 Enum 없이 진행합니다.")

    # ── 2단계: Proto_*.xlsx 파싱 ────────────────────────────────────────
    proto_files = sorted(schema_dir.glob("Proto_*.xlsx"))
    if not proto_files:
        raise FileNotFoundError(
            f"Proto_*.xlsx 파일을 찾을 수 없습니다: {schema_dir}\n"
            "스키마 파일 이름은 반드시 'Proto_'로 시작해야 합니다."
        )

    for proto_file in proto_files:
        print(f"  스키마 파일 파싱: {proto_file.name}")
        wb = _open_workbook(proto_file)

        if "_Schema" not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"{proto_file.name}: _Schema 시트가 없습니다. "
                "Proto_ 파일에는 반드시 _Schema 시트가 있어야 합니다."
            )

        # 스키마 병합
        file_schema = _parse_schema(wb["_Schema"])
        for msg_name, fields in file_schema.items():
            if msg_name in merged_schema:
                raise ValueError(
                    f"MessageName '{msg_name}'이 중복 정의되어 있습니다.\n"
                    f"  파일: {proto_file.name}"
                )
            merged_schema[msg_name] = fields

        # 테이블 전용 Enum 병합 (선택)
        if "_Enum" in wb.sheetnames:
            _merge_enums(merged_enums, _parse_enums(wb["_Enum"]), proto_file.name)

        wb.close()

    return merged_schema, merged_enums


def parse_data_dir(data_dir: Path) -> dict:
    """
    data_dir 내의 모든 *.xlsx 파일(Proto_ 제외)을 파싱하여 데이터를 병합한다.

    각 파일의 시트명이 MessageName으로 사용된다.
    _Schema, _Enum 시트는 무시한다.
    서로 다른 파일에서 동일한 시트명(MessageName)이 중복되면 오류를 발생시킨다.

    Args:
        data_dir: 데이터 *.xlsx 파일들이 위치하는 디렉터리 (하위 폴더 미포함)

    Returns:
        {MessageName: [row_dict, ...]} 형태의 병합된 데이터 dict

    Raises:
        FileNotFoundError: data_dir이 없을 때
        ValueError: 동일한 MessageName이 여러 파일에 존재할 때
    """
    data_dir = Path(data_dir)
    if not data_dir.exists():
        raise FileNotFoundError(f"데이터 디렉터리를 찾을 수 없습니다: {data_dir}")

    # 하위 폴더는 제외하고 현재 디렉터리의 .xlsx만 수집, Proto_ 파일 제외
    data_files = sorted(
        f for f in data_dir.glob("*.xlsx")
        if not f.stem.startswith("Proto_")
    )

    if not data_files:
        raise FileNotFoundError(
            f"데이터 .xlsx 파일을 찾을 수 없습니다: {data_dir}\n"
            "Proto_로 시작하지 않는 .xlsx 파일이 필요합니다."
        )

    reserved = {"_Schema", "_Enum"}
    merged_data: dict[str, list] = {}

    for data_file in data_files:
        print(f"  데이터 파일 파싱: {data_file.name}")
        wb = _open_workbook(data_file)

        for sheet_name in wb.sheetnames:
            if sheet_name in reserved:
                continue
            if sheet_name in merged_data:
                raise ValueError(
                    f"MessageName(시트명) '{sheet_name}'이 중복되어 있습니다.\n"
                    f"  파일: {data_file.name}\n"
                    "  동일한 시트명은 하나의 데이터 파일에만 있어야 합니다."
                )
            merged_data[sheet_name] = _parse_data_sheet(wb[sheet_name], sheet_name)

        wb.close()

    return merged_data


def parse_excel_multi(schema_dir: Path, data_dir: Path) -> ExcelData:
    """
    멀티 파일 모드: Proto_*.xlsx 스키마 파일들과 *.xlsx 데이터 파일들을 통합 파싱한다.

    Args:
        schema_dir: Proto_*.xlsx 파일들이 위치하는 디렉터리
        data_dir:   데이터 *.xlsx 파일들이 위치하는 디렉터리

    Returns:
        ExcelData: 병합된 schema, enums, data를 담은 컨테이너
    """
    schema, enums = parse_schema_dir(Path(schema_dir))
    data = parse_data_dir(Path(data_dir))
    return ExcelData(schema=schema, enums=enums, data=data)
